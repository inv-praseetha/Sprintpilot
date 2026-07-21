import os
import copy
import datetime
import openpyxl
from django.http import FileResponse
from django.conf import settings
from django.db import transaction
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from project.models import Project
from accounts.models import EmployeeProfile
from sprints.models import Sprint, SprintTask
from sprints.serializers import SprintSerializer, SprintTaskSerializer

class SprintDownloadTemplateView(APIView):
    """
    API View to serve the static Excel template file stored on the backend.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'tasks_template.xlsx')
        if not os.path.exists(template_path):
            return Response(
                {"detail": "Template file not found on server."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        response = FileResponse(
            open(template_path, 'rb'), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tasks_template.xlsx"'
        return response

class SprintDownloadScheduleView(APIView):
    """
    API View to load the Excel gantt_template.xlsx, populate the sprint project info and tasks,
    and return the styled Excel sheet with auto-updating formulas and conditional formatting.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, sprint_id, *args, **kwargs):
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        template_path = os.path.join(settings.BASE_DIR, 'templates', 'gantt_template.xlsx')
        if not os.path.exists(template_path):
            return Response(
                {"detail": "Gantt template file not found on server."}, 
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 1. Populate project header info
            ws["B1"] = sprint.project.name
            
            start_date = sprint.start_date
            if isinstance(start_date, str):
                start_date = datetime.date.fromisoformat(start_date)
            
            # Write to E2 (Project Start) and E3 (Display Week) to align template formulas
            ws["E2"] = start_date
            ws["E3"] = 1 # Display Week is normally 1
            
            # 2. Gather tasks and group by category
            tasks = sprint.tasks.all().order_by('created_at')
            
            # Categories in required order
            category_mapping = [
                ('UI', 'UI Development'),
                ('Backend', 'Backend Development'),
                ('INFRA', 'Infra Development'),
                ('QA', 'QA Development')
            ]
            
            # Save original styles from the template rows (B through G)
            styles = {}
            for cat_key, cat_name in category_mapping:
                # Determine template row based on original cell positions
                if cat_name == 'UI Development':
                    template_row = 10
                elif cat_name == 'Backend Development':
                    template_row = 20
                elif cat_name == 'Infra Development':
                    template_row = 30
                else: # QA Development / QA
                    template_row = 35
                    
                styles[cat_name] = []
                for col in range(2, 8): # columns B to G (2 to 7)
                    cell = ws.cell(row=template_row, column=col)
                    styles[cat_name].append({
                        'fill': copy.copy(cell.fill) if cell.fill else None,
                        'font': copy.copy(cell.font) if cell.font else None,
                        'border': copy.copy(cell.border) if cell.border else None,
                        'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                        'number_format': cell.number_format
                    })
                    
            # Also save standard task row style (using Row 11)
            styles['task'] = []
            for col in range(2, 8):
                cell = ws.cell(row=11, column=col)
                styles['task'].append({
                    'fill': copy.copy(cell.fill) if cell.fill else None,
                    'font': copy.copy(cell.font) if cell.font else None,
                    'border': copy.copy(cell.border) if cell.border else None,
                    'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format
                })
                
            # 3. Clear schedule area columns B to G in rows 7 to 51
            for r in range(7, 52):
                for c in range(2, 8):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    # Apply regular task style to clean up any old header formats in that row
                    style_info = styles['task'][c - 2]
                    if style_info['fill']: cell.fill = style_info['fill']
                    if style_info['font']: cell.font = style_info['font']
                    if style_info['border']: cell.border = style_info['border']
                    if style_info['alignment']: cell.alignment = style_info['alignment']
                    cell.number_format = style_info['number_format']
            
            # 4. Write data sequentially starting at Row 10
            current_row = 10
            for cat_key, cat_name in category_mapping:
                cat_tasks = [t for t in tasks if t.category == cat_key]
                
                # Write Phase Header
                style_list = styles[cat_name]
                for col_idx in range(2, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    style_info = style_list[col_idx - 2]
                    if style_info['fill']: cell.fill = style_info['fill']
                    if style_info['font']: cell.font = style_info['font']
                    if style_info['border']: cell.border = style_info['border']
                    if style_info['alignment']: cell.alignment = style_info['alignment']
                    cell.number_format = style_info['number_format']
                
                ws.cell(row=current_row, column=2, value=cat_name) # Column B
                current_row += 1
                
                # Write tasks under this phase
                for task in cat_tasks:
                    # Write values
                    ws.cell(row=current_row, column=2, value=task.title) # Column B: Task Name
                    
                    # Column C: Assigned To
                    assignee_name = ""
                    if task.assigned_employee and task.assigned_employee.user:
                        assignee_name = task.assigned_employee.user.full_name or ""
                    ws.cell(row=current_row, column=3, value=assignee_name)
                    
                    # Column D: Progress
                    progress_val = 0.0
                    if task.status == 'DONE':
                        progress_val = 1.0
                    elif task.status == 'QA':
                        progress_val = 0.90
                    elif task.status == 'IN_REVIEW':
                        progress_val = 0.80
                    elif task.status == 'IN_PROGRESS':
                        progress_val = 0.50
                    ws.cell(row=current_row, column=4, value=progress_val)
                    
                    # Column E: Start Date
                    t_start = task.planned_start_date
                    if isinstance(t_start, str):
                        t_start = datetime.date.fromisoformat(t_start)
                    ws.cell(row=current_row, column=5, value=t_start)
                    
                    # Column F: End Date
                    t_end = task.planned_end_date
                    if isinstance(t_end, str):
                        t_end = datetime.date.fromisoformat(t_end)
                    ws.cell(row=current_row, column=6, value=t_end)
                    
                    # Column G: Remarks
                    rec = task.recommendations.filter(accepted=True).first()
                    if not rec and task.assigned_employee:
                        rec = task.recommendations.filter(recommended_employee=task.assigned_employee).first()
                    
                    remarks_parts = []
                    if rec and rec.reason:
                        remarks_parts.append(rec.reason)
                    else:
                        remarks_parts.append(f"Priority: {task.priority}")
                    if task.jira_id:
                        remarks_parts.append(f"({task.jira_id})")
                    
                    remarks = " ".join(remarks_parts)
                    ws.cell(row=current_row, column=7, value=remarks)
                    
                    current_row += 1

            # Remove the old conditional formatting rule for the Gantt chart area
            keys_to_remove = []
            for key in list(ws.conditional_formatting._cf_rules.keys()):
                if "H7" in key.sqref or "BK51" in key.sqref:
                    keys_to_remove.append(key)
            for key in keys_to_remove:
                del ws.conditional_formatting._cf_rules[key]

            # Re-apply the weekend rule and the Gantt highlight rules to the active task rows (H10:BK{current_row - 1})
            if current_row - 1 >= 10:
                cf_range = f"H10:BK{current_row - 1}"
                
                from openpyxl.formatting.rule import FormulaRule
                from openpyxl.styles import PatternFill
                
                weekend_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
                gantt_fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
                
                weekend_rule = FormulaRule(formula=['OR(TEXT(H$4,"ddd")="Sat", TEXT(H$4,"ddd")="Sun", COUNTIF($B$680:$B$696,H$4)>0)'], fill=weekend_fill)
                gantt_rule_cd = FormulaRule(formula=['AND(H$4>=$C10,H$4<=$D10)'], fill=gantt_fill)
                gantt_rule_ef = FormulaRule(formula=['AND(H$4>=$E10,H$4<=$F10)'], fill=gantt_fill)
                
                ws.conditional_formatting.add(cf_range, weekend_rule)
                ws.conditional_formatting.add(cf_range, gantt_rule_cd)
                ws.conditional_formatting.add(cf_range, gantt_rule_ef)
            
            import io
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            clean_sprint_name = "".join([c if c.isalnum() else "_" for c in sprint.milestone])
            filename = f"Schedule_{clean_sprint_name}.xlsx"
            
            response = FileResponse(
                buffer, 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {"detail": f"Excel generation failed: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class SprintListCreateView(APIView):
    """
    API View to handle listing sprints for a project and creating a sprint with tasks.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id, *args, **kwargs):
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

        sprints = Sprint.objects.filter(project=project).prefetch_related(
            'tasks__assigned_employee__user'
        )
        serializer = SprintSerializer(sprints, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, project_id, *args, **kwargs):
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

        if project.status == 'COMPLETED':
            return Response(
                {"detail": "Cannot create sprints in a completed project."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        data = request.data
        milestone = data.get('milestone') or data.get('name')
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if not milestone or not start_date or not end_date:
            return Response(
                {"detail": "Sprint milestone, start_date, and end_date are required."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                sprint = Sprint.objects.create(
                    project=project,
                    milestone=milestone,
                    start_date=start_date,
                    end_date=end_date,
                    status=data.get('status') or 'PLANNED'
                )

                tasks_data = data.get('tasks') or []
                for task_item in tasks_data:
                    title = task_item.get('title')
                    if not title:
                        continue

                    cat = task_item.get('category', 'UI')
                    cat_upper = str(cat).upper().strip()
                    if cat_upper == 'BACKEND':
                        cat = 'Backend'
                    elif cat_upper == 'INFRA':
                        cat = 'INFRA'
                    elif cat_upper == 'UI':
                        cat = 'UI'
                    elif cat_upper == 'QA':
                        cat = 'QA'
                    else:
                        cat = 'UI'

                    priority = task_item.get('priority', 'Normal')
                    if priority not in ['Low', 'Normal', 'High', 'Critical']:
                        priority = 'Normal'

                    status_val = task_item.get('status', 'TODO')
                    status_val_clean = str(status_val).upper().replace(' ', '_').strip()
                    if status_val_clean == 'IN_PROGRESS':
                        status_val = 'IN_PROGRESS'
                    elif status_val_clean == 'COMPLETED' or status_val_clean == 'DONE':
                        status_val = 'DONE'
                    elif status_val_clean == 'TODO':
                        status_val = 'TODO'
                    elif status_val_clean == 'IN_REVIEW':
                        status_val = 'IN_REVIEW'
                    elif status_val_clean == 'QA':
                        status_val = 'QA'
                    elif status_val_clean == 'BLOCKED':
                        status_val = 'BLOCKED'
                    else:
                        status_val = 'TODO'

                    SprintTask.objects.create(
                        sprint=sprint,
                        title=title,
                        description=task_item.get('description') or task_item.get('desc') or '',
                        category=cat,
                        jira_id=task_item.get('jira_id') or task_item.get('jiraId') or '',
                        priority=priority,
                        status=status_val,
                        story_points=task_item.get('story_points') or task_item.get('storyPoints') or None,
                        estimated_hours=task_item.get('estimated_hours') or task_item.get('estimatedHours') or None,
                        planned_start_date=task_item.get('planned_start_date') or None,
                        planned_end_date=task_item.get('planned_end_date') or None,
                        backlog_task_id=task_item.get('backlog_task_id') or ''
                    )

            serializer = SprintSerializer(sprint)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class SprintDetailView(APIView):
    """
    API View to retrieve or delete a specific sprint.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk, *args, **kwargs):
        try:
            sprint = Sprint.objects.prefetch_related('tasks__assigned_employee__user').get(id=pk)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = SprintSerializer(sprint)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk, *args, **kwargs):
        try:
            sprint = Sprint.objects.get(id=pk)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        sprint.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class SprintTaskUpdateView(APIView):
    """
    API View to update single fields of a sprint task.
    """
    permission_classes = [IsAuthenticated]

    def put(self, request, pk, *args, **kwargs):
        try:
            task = SprintTask.objects.get(id=pk)
        except SprintTask.DoesNotExist:
            return Response({"detail": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

        if task.sprint.project.status == 'COMPLETED':
            return Response(
                {"detail": "Cannot update tasks in a completed project."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        data = request.data

        if 'status' in data:
            status_val = data['status']
            status_val_clean = str(status_val).upper().replace(' ', '_').strip()
            if status_val_clean == 'IN_PROGRESS':
                task.status = 'IN_PROGRESS'
            elif status_val_clean in ('COMPLETED', 'DONE', 'CLOSED'):
                task.status = 'CLOSED'
            elif status_val_clean in ('TODO', 'OPEN'):
                task.status = 'OPEN'
            elif status_val_clean in ('IN_REVIEW', 'QA', 'RESOLVED'):
                task.status = 'RESOLVED'
            else:
                task.status = status_val

        if 'assigned_employee_id' in data or 'assignedTo' in data:
            emp_id = data.get('assigned_employee_id') or data.get('assignedTo')
            if emp_id:
                if emp_id == 'unassigned' or emp_id == '':
                    task.assigned_employee = None
                else:
                    try:
                        task.assigned_employee = EmployeeProfile.objects.get(id=emp_id)
                    except EmployeeProfile.DoesNotExist:
                        return Response({"detail": "Employee profile not found."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                task.assigned_employee = None

        if 'planned_start_date' in data or 'startDate' in data:
            task.planned_start_date = data.get('planned_start_date') or data.get('startDate')

        if 'planned_end_date' in data or 'endDate' in data:
            task.planned_end_date = data.get('planned_end_date') or data.get('endDate')

        try:
            task.save()
            serializer = SprintTaskSerializer(task)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# View endpoints delegated to services
class SprintAISuggestScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, sprint_id, *args, **kwargs):
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        if sprint.project.status == 'COMPLETED':
            return Response(
                {"detail": "Cannot generate AI schedule for a completed project."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        from decouple import config
        api_key = config('GEMINI_API_KEY', default=None) or os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return Response(
                {"detail": "GEMINI_API_KEY is not configured. Please add GEMINI_API_KEY to your backend .env file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        task_ids = request.data.get('task_ids', [])
        if task_ids:
            tasks = sprint.tasks.filter(id__in=task_ids)
        else:
            tasks = sprint.tasks.all()

        if not tasks.exists():
            return Response(
                {"detail": "No tasks found in this sprint to schedule."},
                status=status.HTTP_400_BAD_REQUEST
            )

        from sprints.services.schedule_service import generate_and_persist_recommendations
        try:
            output_suggestions = generate_and_persist_recommendations(sprint, tasks, api_key)
            return Response(output_suggestions, status=status.HTTP_200_OK)
        except ImportError as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {"detail": f"AI Generation Failed: {str(e)}"},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

class SprintImportScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, sprint_id, *args, **kwargs):
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        if sprint.project.status == 'COMPLETED':
            return Response(
                {"detail": "Cannot import schedule for a completed project."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        from sprints.services.schedule_service import import_schedule
        try:
            import_schedule(sprint, request.data)
            return Response({"detail": "Schedule successfully imported and saved."}, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SprintTaskCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, sprint_id, *args, **kwargs):
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            return Response({"detail": "Sprint not found."}, status=status.HTTP_404_NOT_FOUND)

        if sprint.project.status == 'COMPLETED':
            return Response(
                {"detail": "Cannot add tasks to a completed project."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = SprintTaskSerializer(data=request.data)
        if serializer.is_valid():
            task = serializer.save(sprint=sprint)
            
            # Recalculate story points and hours if planned start and end date are provided
            start_date = serializer.validated_data.get('planned_start_date')
            end_date = serializer.validated_data.get('planned_end_date')
            if start_date and end_date:
                from sprints.services.schedule_service import calculate_working_days
                start_str = start_date.strftime("%Y-%m-%d")
                end_str = end_date.strftime("%Y-%m-%d")
                wd = calculate_working_days(start_str, end_str)
                task.story_points = wd * 2
                task.estimated_hours = wd * 8
                task.save()
            
            # Update TaskRecommendation states if this task matches the assigned employee
            if task.assigned_employee:
                from sprints.models import TaskRecommendation
                TaskRecommendation.objects.filter(task=task, recommended_employee=task.assigned_employee).update(accepted=True)
                TaskRecommendation.objects.filter(task=task).exclude(recommended_employee=task.assigned_employee).update(accepted=False)
                
            return Response(SprintTaskSerializer(task).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

