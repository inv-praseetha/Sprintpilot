import os
import io
import copy
import datetime
import openpyxl
from django.conf import settings
from django.db import transaction
from django.db.models import Prefetch

from project.models import Project
from accounts.models import EmployeeProfile
from sprints.models import Sprint, SprintTask, SprintHoliday, TaskRecommendation, SprintNote
from sprints.serializers import SprintSerializer, SprintTaskSerializer


class SprintService:

    @staticmethod
    def get_template_path() -> str:
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'tasks_template.xlsx')
        if not os.path.exists(template_path):
            raise FileNotFoundError("Template file not found on server.")
        return template_path

    @staticmethod
    def generate_excel_schedule(sprint_id: str) -> tuple[io.BytesIO, str]:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        template_path = os.path.join(settings.BASE_DIR, 'templates', 'gantt_template.xlsx')
        if not os.path.exists(template_path):
            raise FileNotFoundError("Gantt template file not found on server.")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 1. Populate project header info
        ws["B1"] = sprint.project.name
        
        start_date = sprint.start_date
        if isinstance(start_date, str):
            start_date = datetime.date.fromisoformat(start_date)
        
        # Write to E2 (Project Start) and E3 (Display Week) to align template formulas
        ws["E2"] = start_date
        ws["E3"] = 1 # Display Week is normally 1
        
        # 2. Gather tasks and group by category
        tasks = sprint.tasks.filter(is_deleted=False).order_by('created_at')
        
        # Categories in required order
        category_mapping = [
            ('UI', 'UI Development'),
            ('Backend', 'Backend Development'),
            ('INFRA', 'Infra Development'),
            ('QA', 'QA Development')
        ]
        
        # Save original styles from the template rows (B through G)
        styles = {}
        for cat_key, cat_name in category_mapping:
            # Determine template row based on original cell positions
            if cat_name == 'UI Development':
                template_row = 10
            elif cat_name == 'Backend Development':
                template_row = 20
            elif cat_name == 'Infra Development':
                template_row = 30
            else: # QA Development / QA
                template_row = 35
                
            styles[cat_name] = []
            for col in range(2, 8): # columns B to G (2 to 7)
                cell = ws.cell(row=template_row, column=col)
                styles[cat_name].append({
                    'fill': copy.copy(cell.fill) if cell.fill else None,
                    'font': copy.copy(cell.font) if cell.font else None,
                    'border': copy.copy(cell.border) if cell.border else None,
                    'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format
                })
                
        # Save task styles per category
        # UI task template: row 11
        styles['task_UI'] = []
        for col in range(2, 8):
            cell = ws.cell(row=11, column=col)
            styles['task_UI'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })

        # Backend task template: row 21
        styles['task_Backend'] = []
        for col in range(2, 8):
            cell = ws.cell(row=21, column=col)
            styles['task_Backend'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })

        # Infra task template: row 31
        styles['task_INFRA'] = []
        for col in range(2, 8):
            cell = ws.cell(row=31, column=col)
            styles['task_INFRA'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })

        # QA task template: row 36
        styles['task_QA'] = []
        for col in range(2, 8):
            cell = ws.cell(row=36, column=col)
            styles['task_QA'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })
            
        # Save Releases row style (using Row 49)
        styles['Releases'] = []
        for col in range(2, 8):
            cell = ws.cell(row=49, column=col)
            styles['Releases'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })

        # Save UAT Release row style (using Row 50)
        styles['UAT_Release'] = []
        for col in range(2, 8):
            cell = ws.cell(row=50, column=col)
            styles['UAT_Release'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })

        # Save Production Release row style (using Row 51)
        styles['Production_Release'] = []
        for col in range(2, 8):
            cell = ws.cell(row=51, column=col)
            styles['Production_Release'].append({
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            })
            
        # 3. Clear schedule area columns B to G in rows 7 to 51
        for r in range(7, 52):
            for c in range(2, 8):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                # Apply regular task style to clean up any old header formats in that row
                style_info = styles['task_UI'][c - 2]
                if style_info['fill']: cell.fill = style_info['fill']
                if style_info['font']: cell.font = style_info['font']
                if style_info['border']: cell.border = style_info['border']
                if style_info['alignment']: cell.alignment = style_info['alignment']
                cell.number_format = style_info['number_format']
        
        # 4. Write data sequentially starting at Row 10
        current_row = 10
        for cat_key, cat_name in category_mapping:
            cat_tasks = [t for t in tasks if t.category == cat_key]
            
            # Write Phase Header
            style_list = styles[cat_name]
            for col_idx in range(2, 8):
                cell = ws.cell(row=current_row, column=col_idx)
                style_info = style_list[col_idx - 2]
                if style_info['fill']: cell.fill = style_info['fill']
                if style_info['font']: cell.font = style_info['font']
                if style_info['border']: cell.border = style_info['border']
                if style_info['alignment']: cell.alignment = style_info['alignment']
                cell.number_format = style_info['number_format']
            
            ws.cell(row=current_row, column=2, value=cat_name) # Column B
            current_row += 1
            
            # Write tasks under this phase
            for task in cat_tasks:
                # Apply task style for this category
                style_list = styles[f'task_{cat_key}']
                for col_idx in range(2, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    style_info = style_list[col_idx - 2]
                    if style_info['fill']: cell.fill = style_info['fill']
                    if style_info['font']: cell.font = style_info['font']
                    if style_info['border']: cell.border = style_info['border']
                    if style_info['alignment']: cell.alignment = style_info['alignment']
                    cell.number_format = style_info['number_format']

                # Write values
                cell_b = ws.cell(row=current_row, column=2)
                if task.jira_id:
                    # Dynamically get the workspace URL from the authenticated Jira session
                    from jira_integration.models import JiraOAuthToken
                    jira_token = JiraOAuthToken.objects.first()
                    jira_base = jira_token.workspace_url if (jira_token and jira_token.workspace_url) else "https://jira.atlassian.com"
                    
                    clean_base = jira_base.rstrip('/')
                    
                    if task.jira_id.startswith('http'):
                        jira_url = task.jira_id
                    else:
                        # Use the canonical Jira Cloud URL format which automatically handles redirects
                        jira_url = f"{clean_base}/browse/{task.jira_id}"
                        
                    # Use a very safe Excel formula. 
                    jira_url = jira_url.strip()
                    safe_title = task.title.replace('"', "''").replace('\n', ' ')
                    
                    # Calculate safe display length to keep the ENTIRE formula under 255 characters
                    # Formula: =HYPERLINK("url", "display")
                    # Formula overhead: 16 chars + len(jira_url) + len(display_text)
                    max_display_len = 250 - 16 - len(jira_url)
                    
                    display_text = f"{safe_title}"
                    if len(display_text) > max_display_len:
                        display_text = display_text[:max_display_len-3] + "..."
                        
                    cell_b.value = f'=HYPERLINK("{jira_url}", "{display_text}")'
                    
                    # Style the font to look like a hyperlink
                    cell_b.font = openpyxl.styles.Font(
                        name=cell_b.font.name if cell_b.font else 'Segoe UI',
                        size=cell_b.font.size if cell_b.font else 10,
                        bold=cell_b.font.bold if cell_b.font else False,
                        italic=cell_b.font.italic if cell_b.font else False,
                        underline='single',
                        color='0563C1'
                    )
                    
                    # Enable wrap text
                    orig_align = cell_b.alignment
                    cell_b.alignment = openpyxl.styles.Alignment(
                        horizontal=orig_align.horizontal if orig_align else None,
                        vertical=orig_align.vertical if orig_align else None,
                        text_rotation=orig_align.text_rotation if orig_align else 0,
                        wrap_text=True,
                        shrink_to_fit=orig_align.shrink_to_fit if orig_align else False,
                        indent=orig_align.indent if orig_align else 0
                    )
                    # Set height to accommodate two lines
                    ws.row_dimensions[current_row].height = 28
                else:
                    cell_b.value = task.title
                
                # Column C: Assigned To
                assignee_name = ""
                if task.assigned_employee and task.assigned_employee.user:
                    assignee_name = task.assigned_employee.user.full_name or ""
                ws.cell(row=current_row, column=3, value=assignee_name)
                
                # Column D: Progress
                progress_val = 0.0
                if task.status == 'DONE':
                    progress_val = 1.0
                elif task.status == 'QA':
                    progress_val = 0.90
                elif task.status == 'IN_REVIEW':
                    progress_val = 0.80
                elif task.status == 'IN_PROGRESS':
                    progress_val = 0.50
                ws.cell(row=current_row, column=4, value=progress_val)
                
                # Column E: Start Date
                t_start = task.planned_start_date
                if isinstance(t_start, str):
                    t_start = datetime.date.fromisoformat(t_start)
                ws.cell(row=current_row, column=5, value=t_start)
                
                # Column F: End Date
                t_end = task.planned_end_date
                if isinstance(t_end, str):
                    t_end = datetime.date.fromisoformat(t_end)
                ws.cell(row=current_row, column=6, value=t_end)
                
                # Column G: Remarks
                ws.cell(row=current_row, column=7, value=None)
                
                current_row += 1

        # Write "Releases" header row (just color and name, no tasks, no remarks)
        style_list = styles['Releases']
        for col_idx in range(2, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            style_info = style_list[col_idx - 2]
            if style_info['fill']: cell.fill = style_info['fill']
            if style_info['font']: cell.font = style_info['font']
            if style_info['border']: cell.border = style_info['border']
            if style_info['alignment']: cell.alignment = style_info['alignment']
            cell.number_format = style_info['number_format']
        
        ws.cell(row=current_row, column=2, value='Releases') # Column B
        # Ensure other columns in this row are empty (no remarks etc.)
        for col_idx in range(3, 8):
            ws.cell(row=current_row, column=col_idx, value=None)
        
        current_row += 1

        # Write "UAT Release" task row
        sprint_end_dt = sprint.end_date
        if isinstance(sprint_end_dt, str):
            sprint_end_dt = datetime.date.fromisoformat(sprint_end_dt)
        second_last_day = sprint_end_dt - datetime.timedelta(days=1)
        last_day = sprint_end_dt

        style_list = styles['UAT_Release']
        for col_idx in range(2, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            style_info = style_list[col_idx - 2]
            if style_info['fill']: cell.fill = style_info['fill']
            if style_info['font']: cell.font = style_info['font']
            if style_info['border']: cell.border = style_info['border']
            if style_info['alignment']: cell.alignment = style_info['alignment']
            cell.number_format = style_info['number_format']
        
        ws.cell(row=current_row, column=2, value='UAT Release')
        ws.cell(row=current_row, column=3, value=None) # Assigned To: leave empty
        ws.cell(row=current_row, column=4, value=0.0)  # Progress: 0%
        ws.cell(row=current_row, column=5, value=second_last_day) # Start Date
        ws.cell(row=current_row, column=6, value=second_last_day) # End Date
        ws.cell(row=current_row, column=7, value=None) # Remarks
        
        current_row += 1

        # Write "Production Release" task row
        style_list = styles['Production_Release']
        for col_idx in range(2, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            style_info = style_list[col_idx - 2]
            if style_info['fill']: cell.fill = style_info['fill']
            if style_info['font']: cell.font = style_info['font']
            if style_info['border']: cell.border = style_info['border']
            if style_info['alignment']: cell.alignment = style_info['alignment']
            cell.number_format = style_info['number_format']
        
        ws.cell(row=current_row, column=2, value='Production Release')
        ws.cell(row=current_row, column=3, value=None) # Assigned To: leave empty
        ws.cell(row=current_row, column=4, value=0.0)  # Progress: 0%
        ws.cell(row=current_row, column=5, value=last_day) # Start Date
        ws.cell(row=current_row, column=6, value=last_day) # End Date
        ws.cell(row=current_row, column=7, value=None) # Remarks
        
        current_row += 1

        # Remove the old conditional formatting rule for the Gantt chart area
        keys_to_remove = []
        for key in list(ws.conditional_formatting._cf_rules.keys()):
            if "H7" in key.sqref or "BK51" in key.sqref:
                keys_to_remove.append(key)
        for key in keys_to_remove:
            del ws.conditional_formatting._cf_rules[key]

        # Re-apply the weekend rule and the Gantt highlight rules to the active task rows (H10:BK{current_row - 1})
        if current_row - 1 >= 10:
            cf_range = f"H10:BK{current_row - 1}"
            
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import PatternFill
            
            weekend_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            gantt_fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
            
            weekend_rule = FormulaRule(formula=['OR(TEXT(H$4,"ddd")="Sat", TEXT(H$4,"ddd")="Sun", COUNTIF($B$680:$B$696,H$4)>0)'], fill=weekend_fill)
            gantt_rule_cd = FormulaRule(formula=['AND(H$4>=$C10,H$4<=$D10)'], fill=gantt_fill)
            gantt_rule_ef = FormulaRule(formula=['AND(H$4>=$E10,H$4<=$F10)'], fill=gantt_fill)
            
            ws.conditional_formatting.add(cf_range, weekend_rule)
            ws.conditional_formatting.add(cf_range, gantt_rule_cd)
            ws.conditional_formatting.add(cf_range, gantt_rule_ef)
        
        # Set Column E (Start Date) width to match Column F (End Date) width
        ws.column_dimensions['E'].width = ws.column_dimensions['F'].width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        clean_sprint_name = "".join([c if c.isalnum() else "_" for c in sprint.milestone])
        filename = f"Schedule_{clean_sprint_name}.xlsx"
        return buffer, filename

    @staticmethod
    def list_sprints(project_id: str):
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            raise Project.DoesNotExist("Project not found.")

        sprints = Sprint.objects.filter(project=project, is_deleted=False).select_related('project').prefetch_related(
            Prefetch('tasks', queryset=SprintTask.objects.filter(is_deleted=False)),
            'tasks__assigned_employee__user',
            'tasks__assigned_employee__employee_skill_relations__skill',
            'tasks__recommendations'
        )
        return sprints

    @staticmethod
    def create_sprint(project_id: str, data: dict) -> Sprint:
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            raise Project.DoesNotExist("Project not found.")

        if project.status == 'COMPLETED':
            raise ValueError("Cannot create sprints in a completed project.")

        milestone = data.get('milestone') or data.get('name')
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if not milestone or not start_date or not end_date:
            raise ValueError("Sprint milestone, start_date, and end_date are required.")

        # Parse and validate sprint dates
        try:
            if isinstance(start_date, str):
                parsed_start = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                parsed_start = start_date
            if isinstance(end_date, str):
                parsed_end = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                parsed_end = end_date
                
            if parsed_start > parsed_end:
                raise ValueError("Sprint start date must be before or equal to end date.")
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid date formats: {str(e)}")

        with transaction.atomic():
            sprint = Sprint.objects.create(
                project=project,
                milestone=milestone,
                start_date=start_date,
                end_date=end_date,
                status=data.get('status') or 'ACTIVE',
                backlog_version_id=data.get('backlog_version_id')
            )

            holidays_data = data.get('holidays') or []
            for holiday_str in holidays_data:
                try:
                    h_date = datetime.datetime.strptime(holiday_str, "%Y-%m-%d").date()
                    SprintHoliday.objects.get_or_create(
                        sprint=sprint,
                        date=h_date,
                        defaults={'description': 'Sprint Holiday'}
                    )
                except Exception:
                    pass

            tasks_data = data.get('tasks') or []
            for task_item in tasks_data:
                title = task_item.get('title')
                if not title:
                    continue

                cat_val = task_item.get('category', 'UI')
                if isinstance(cat_val, list):
                    cat = ", ".join(cat_val)
                else:
                    cat = str(cat_val).strip()

                priority = task_item.get('priority', 'Normal')
                if priority not in ['Low', 'Normal', 'High', 'Critical']:
                    priority = 'Normal'

                status_val = task_item.get('status', 'OPEN')
                status_val_clean = str(status_val).upper().replace(' ', '_').strip()
                if status_val_clean == 'IN_PROGRESS':
                    status_val = 'IN_PROGRESS'
                elif status_val_clean in ('COMPLETED', 'DONE', 'CLOSED'):
                    status_val = 'CLOSED'
                elif status_val_clean in ('TODO', 'OPEN'):
                    status_val = 'OPEN'
                elif status_val_clean in ('IN_REVIEW', 'QA', 'RESOLVED'):
                    status_val = 'RESOLVED'
                else:
                    status_val = 'OPEN'

                SprintTask.objects.create(
                    sprint=sprint,
                    title=title,
                    description=task_item.get('description') or task_item.get('desc') or '',
                    category=cat,
                    jira_id=task_item.get('jira_id') or task_item.get('jiraId') or '',
                    priority=priority,
                    status=status_val,
                    story_points=task_item.get('story_points') or task_item.get('storyPoints') or None,
                    estimated_hours=task_item.get('estimated_hours') or task_item.get('estimatedHours') or None,
                    planned_start_date=task_item.get('planned_start_date') or None,
                    planned_end_date=task_item.get('planned_end_date') or None,
                    backlog_task_id=task_item.get('backlog_task_id') or ''
                )

        return sprint

    @staticmethod
    def get_sprint_detail(sprint_id: str) -> Sprint:
        try:
            sprint = Sprint.objects.select_related('project').prefetch_related(
                Prefetch('tasks', queryset=SprintTask.objects.filter(is_deleted=False)),
                'tasks__assigned_employee__user',
                'tasks__assigned_employee__employee_skill_relations__skill',
                'tasks__recommendations'
            ).get(id=sprint_id, is_deleted=False)
            return sprint
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

    @staticmethod
    def delete_sprint(sprint_id: str) -> None:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
            # Soft delete the sprint
            sprint.is_deleted = True
            sprint.save(update_fields=['is_deleted'])
            
            # Soft delete tasks that are NOT synced to backlog
            from django.db.models import Q
            sprint.tasks.filter(Q(backlog_task_id__isnull=True) | Q(backlog_task_id__exact='')).update(is_deleted=True)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

    @staticmethod
    def get_sprint_closure_summary(sprint_id: str) -> dict:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        if sprint.status == 'COMPLETED':
            raise ValueError("Sprint is already closed.")

        if not sprint.synced_at:
            raise ValueError("Sprint is not connected to Backlog. Please sync tasks to Backlog first.")

        tasks = sprint.tasks.filter(is_deleted=False)
        total_tasks = tasks.count()
        if total_tasks == 0:
            raise ValueError("Sprint does not contain any tasks.")

        summary = {
            "total_tasks": total_tasks,
            "open_tasks": tasks.filter(status='OPEN').count(),
            "in_progress_tasks": tasks.filter(status='IN_PROGRESS').count(),
            "resolved_tasks": tasks.filter(status='RESOLVED').count(),
            "closed_tasks": tasks.filter(status='CLOSED').count()
        }
        return summary

    @staticmethod
    def close_sprint(sprint_id: str) -> None:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        if sprint.status == 'COMPLETED':
            raise ValueError("Sprint is already closed.")

        if not sprint.synced_at:
            raise ValueError("Sprint is not connected to Backlog. Please sync tasks to Backlog first.")

        with transaction.atomic():
            tasks = sprint.tasks.filter(is_deleted=False)
            if not tasks.exists():
                raise ValueError("Sprint does not contain any tasks.")
                
            for task in tasks:
                if task.status != 'CLOSED':
                    task.status = 'CLOSED'
                    task._skip_sync_validation = True
                    task.save()
                    
                    if task.backlog_task_id and sprint.project.project_id:
                        try:
                            from backlog.services.backlog_client import BacklogService
                            backlog_client = BacklogService(project_key=sprint.project.project_id)
                            backlog_client.update_task(task)
                        except Exception as e:
                            if "NO_CHANGES_DETECTED" not in str(e):
                                import logging
                                logging.getLogger(__name__).error(f"Failed to close task {task.id} in backlog: {e}")
                                raise Exception(f"Failed to sync task {task.id} to Backlog. Operation aborted.")

            sprint.status = 'COMPLETED'
            sprint.save()

    @staticmethod
    def update_task(task_id: str, data: dict) -> SprintTask:
        try:
            task = SprintTask.objects.get(id=task_id)
        except SprintTask.DoesNotExist:
            raise SprintTask.DoesNotExist("Task not found.")

        if task.sprint.project.status == 'COMPLETED':
            raise ValueError("Cannot update tasks in a completed project.")

        if 'status' in data:
            status_val = data['status']
            status_val_clean = str(status_val).upper().replace(' ', '_').strip()
            if status_val_clean == 'IN_PROGRESS':
                task.status = 'IN_PROGRESS'
            elif status_val_clean in ('COMPLETED', 'DONE', 'CLOSED'):
                task.status = 'CLOSED'
            elif status_val_clean in ('TODO', 'OPEN'):
                task.status = 'OPEN'
            elif status_val_clean in ('IN_REVIEW', 'QA', 'RESOLVED'):
                task.status = 'RESOLVED'
            else:
                task.status = status_val

        if 'assigned_employee_id' in data or 'assignedTo' in data:
            emp_id = data.get('assigned_employee_id') or data.get('assignedTo')
            if emp_id:
                if emp_id == 'unassigned' or emp_id == '':
                    task.assigned_employee = None
                else:
                    try:
                        task.assigned_employee = EmployeeProfile.objects.get(id=emp_id)
                    except EmployeeProfile.DoesNotExist:
                        raise ValueError("Employee profile not found.")
            else:
                task.assigned_employee = None

        # Validate dates and values if updated
        new_start = data.get('planned_start_date') or data.get('startDate')
        new_end = data.get('planned_end_date') or data.get('endDate')
        
        chk_start = new_start if ('planned_start_date' in data or 'startDate' in data) else task.planned_start_date
        chk_end = new_end if ('planned_end_date' in data or 'endDate' in data) else task.planned_end_date
        
        import datetime
        if isinstance(chk_start, str):
            try:
                chk_start = datetime.datetime.strptime(chk_start, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError("Invalid start date format. Must be YYYY-MM-DD.")
        if isinstance(chk_end, str):
            try:
                chk_end = datetime.datetime.strptime(chk_end, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError("Invalid end date format. Must be YYYY-MM-DD.")
                
        if chk_start and chk_end:
            if chk_start > chk_end:
                raise ValueError("Planned start date must be before or equal to planned end date.")
            if task.sprint:
                if chk_start < task.sprint.start_date or chk_end > task.sprint.end_date:
                    raise ValueError(f"Task dates must fall within the sprint range ({task.sprint.start_date} to {task.sprint.end_date}).")
                    
        new_hours = data.get('estimated_hours') if 'estimated_hours' in data else data.get('estimatedHours')
        chk_hours = new_hours if new_hours is not None else task.estimated_hours
        
        if chk_hours is not None:
            try:
                hours_val = float(chk_hours)
                if hours_val < 0:
                    raise ValueError("Estimated hours cannot be negative.")
                if chk_start and chk_end and hours_val > 0:
                    from math import ceil
                    working_days = 0
                    curr = chk_start
                    while curr <= chk_end:
                        if curr.weekday() < 5:
                            working_days += 1
                        curr += datetime.timedelta(days=1)
                    min_days = ceil(hours_val / 8.0)
                    if working_days < min_days:
                        raise ValueError(f"Estimated hours ({hours_val}h) require at least {min_days} working day(s). Selected range has only {working_days} working day(s).")
            except (ValueError, TypeError) as e:
                if "require at least" in str(e):
                    raise ValueError(str(e))
                raise ValueError("Estimated hours must be a valid number.")

        start_changed = 'planned_start_date' in data or 'startDate' in data
        end_changed = 'planned_end_date' in data or 'endDate' in data
        hours_changed = 'estimated_hours' in data or 'estimatedHours' in data
        sp_changed = 'story_points' in data or 'storyPoints' in data
        
        if start_changed:
            task.planned_start_date = data.get('planned_start_date') or data.get('startDate')

        if end_changed:
            task.planned_end_date = data.get('planned_end_date') or data.get('endDate')

        if hours_changed:
            task.estimated_hours = data.get('estimated_hours') if 'estimated_hours' in data else data.get('estimatedHours')
        elif start_changed or end_changed:
            if task.planned_start_date and task.planned_end_date:
                from sprints.services.schedule_service import calculate_working_days
                holidays = list(task.sprint.holidays.values_list('date', flat=True)) if task.sprint else None
                wd = calculate_working_days(str(task.planned_start_date), str(task.planned_end_date), holidays=holidays)
                if task.estimated_hours is None:
                    task.estimated_hours = wd * 8
            
        if sp_changed:
            task.story_points = data.get('story_points') if 'story_points' in data else data.get('storyPoints')
        elif start_changed or end_changed:
            if task.planned_start_date and task.planned_end_date:
                from sprints.services.schedule_service import calculate_working_days
                holidays = list(task.sprint.holidays.values_list('date', flat=True)) if task.sprint else None
                wd = calculate_working_days(str(task.planned_start_date), str(task.planned_end_date), holidays=holidays)
                if task.story_points is None:
                    task.story_points = wd * 2
            
        if 'title' in data:
            task.title = data.get('title')
            
        if 'description' in data:
            task.description = data.get('description')
            
        if 'priority' in data:
            task.priority = data.get('priority')

        task.save()
        return task

    @staticmethod
    def delete_task(task_id: str) -> None:
        try:
            task = SprintTask.objects.get(id=task_id)
        except SprintTask.DoesNotExist:
            raise SprintTask.DoesNotExist("Task not found.")

        if task.sprint.project.status == 'COMPLETED':
            raise ValueError("Cannot delete tasks in a completed project.")

        if task.status == 'CLOSED':
            raise ValueError("Cannot delete this task as it is already closed/completed.")

        if task.backlog_task_id:
            from backlog.services.backlog_client import BacklogService
            BacklogService(project_key=task.sprint.project.project_id).delete_issue(task.backlog_task_id)

        task.is_deleted = True
        task.save()

    @staticmethod
    def bulk_delete_tasks(task_ids: list) -> int:
        if not task_ids:
            raise ValueError("No task IDs provided.")

        tasks = SprintTask.objects.filter(id__in=task_ids)
        if tasks.filter(sprint__project__status='COMPLETED').exists():
            raise ValueError("Cannot delete tasks in a completed project.")

        if tasks.filter(status='CLOSED').exists():
            raise ValueError("Cannot delete this task as it is already closed/completed.")

        count = tasks.count()
        tasks.update(is_deleted=True)
        return count

    @staticmethod
    def generate_ai_suggestions(sprint_id: str, task_ids: list, api_key: str) -> list:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        if sprint.project.status == 'COMPLETED':
            raise ValueError("Cannot generate AI schedule for a completed project.")

        if not api_key:
            raise ValueError("GEMINI_API_KEY is not configured.")

        if task_ids:
            tasks = sprint.tasks.filter(id__in=task_ids, is_deleted=False)
        else:
            tasks = sprint.tasks.filter(is_deleted=False)

        if not tasks.exists():
            raise ValueError("No tasks found in this sprint to schedule.")

        from sprints.services.schedule_service import generate_and_persist_recommendations
        return generate_and_persist_recommendations(sprint, tasks, api_key)

    @staticmethod
    def import_schedule_data(sprint_id: str, data: list) -> None:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        if sprint.project.status == 'COMPLETED':
            raise ValueError("Cannot import schedule for a completed project.")

        from sprints.services.schedule_service import import_schedule
        import_schedule(sprint, data)

    @staticmethod
    def create_task(sprint_id: str, data: dict) -> SprintTask:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        if sprint.project.status == 'COMPLETED':
            raise ValueError("Cannot add tasks to a completed project.")

        serializer = SprintTaskSerializer(data=data, context={'sprint': sprint})
        if not serializer.is_valid():
            raise ValueError(serializer.errors)

        task = serializer.save(sprint=sprint)
        
        # Recalculate story points and hours if planned start and end date are provided
        start_date = serializer.validated_data.get('planned_start_date')
        end_date = serializer.validated_data.get('planned_end_date')
        if start_date and end_date:
            from sprints.services.schedule_service import calculate_working_days
            start_str = start_date.strftime("%Y-%m-%d")
            end_str = end_date.strftime("%Y-%m-%d")
            holidays = list(sprint.holidays.values_list('date', flat=True)) if sprint else None
            wd = calculate_working_days(start_str, end_str, holidays=holidays)
            if task.story_points is None:
                task.story_points = wd * 2
            if task.estimated_hours is None:
                task.estimated_hours = wd * 8
            task.save()
        
        # Update TaskRecommendation states if this task matches the assigned employee
        if task.assigned_employee:
            TaskRecommendation.objects.filter(task=task, recommended_employee=task.assigned_employee).update(accepted=True)
            TaskRecommendation.objects.filter(task=task).exclude(recommended_employee=task.assigned_employee).update(accepted=False)

        return task

    @staticmethod
    def list_sprint_notes(sprint_id: str):
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")
        return SprintNote.objects.filter(sprint=sprint)

    @staticmethod
    def save_sprint_note(sprint_id: str, date_str: str, content: str = None, attachment=None, delete_attachment: bool = False) -> SprintNote:
        try:
            sprint = Sprint.objects.get(id=sprint_id)
        except Sprint.DoesNotExist:
            raise Sprint.DoesNotExist("Sprint not found.")

        try:
            note_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            note_date = datetime.date.today()

        defaults = {}
        if content is not None:
            defaults['content'] = content
        if attachment is not None:
            defaults['attachment'] = attachment
        elif delete_attachment:
            defaults['attachment'] = None

        note, created = SprintNote.objects.update_or_create(
            sprint=sprint,
            date=note_date,
            defaults=defaults
        )
        return note

    @staticmethod
    def get_tasks_by_due_status() -> dict:
        """
        Returns per-project task counts grouped by urgency bucket:
        overdue (past due or no date), due today, due tomorrow.
        Only considers non-closed/resolved tasks in non-completed projects.
        """
        from django.utils import timezone
        from datetime import timedelta
        from project.models import Project

        today = timezone.localdate()
        tomorrow = today + timedelta(days=1)

        active_projects = Project.objects.exclude(status='COMPLETED')

        tasks = SprintTask.objects.filter(
            sprint__project__in=active_projects,
            is_deleted=False
        ).exclude(status__in=['CLOSED', 'RESOLVED']).select_related(
            'sprint__project'
        )

        # project_id -> { projectId, projectName, overdue, today, tomorrow }
        project_map = {}

        for t in tasks:
            pid = str(t.sprint.project.id)
            if pid not in project_map:
                project_map[pid] = {
                    'projectId': pid,
                    'projectName': t.sprint.project.name,
                    'overdue': 0,
                    'today': 0,
                    'tomorrow': 0,
                }

            if not t.planned_end_date or t.planned_end_date < today:
                project_map[pid]['overdue'] += 1
            elif t.planned_end_date == today:
                project_map[pid]['today'] += 1
            elif t.planned_end_date == tomorrow:
                project_map[pid]['tomorrow'] += 1

        projects = list(project_map.values())

        def bucket(key):
            return sorted(
                [{'projectId': p['projectId'], 'projectName': p['projectName'], 'count': p[key]}
                 for p in projects if p[key] > 0],
                key=lambda x: -x['count']
            )

        return {
            'overdue':   bucket('overdue'),
            'today':     bucket('today'),
            'tomorrow':  bucket('tomorrow'),
        }


